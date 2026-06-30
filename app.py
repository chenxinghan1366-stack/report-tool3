{\rtf1\ansi\ansicpg936\cocoartf2903
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fswiss\fcharset0 Helvetica;}
{\colortbl;\red255\green255\blue255;}
{\*\expandedcolortbl;;}
\paperw11900\paperh16840\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\pard\tx720\tx1440\tx2160\tx2880\tx3600\tx4320\tx5040\tx5760\tx6480\tx7200\tx7920\tx8640\pardirnatural\partightenfactor0

\f0\fs24 \cf0 import streamlit as st\
import pandas as pd\
import json\
from datetime import datetime\
from openpyxl import Workbook\
from io import BytesIO\
\
# ---------- \uc0\u27169 \u25311 \u30693 \u35782 \u24211  ----------\
SOURCES = [\
    \{"\uc0\u30465 \u20221 ": "\u19978 \u28023 ", "\u22478 \u24066 ": "\u19978 \u28023 \u24066 ", "\u26426 \u26500 ": "\u19978 \u28023 \u24066 \u31246 \u21153 \u23616 ", "\u26469 \u28304 ": "https://shanghai.chinatax.gov.cn/", "\u32423 \u21035 ": "\u30465 \u32423 "\},\
    \{"\uc0\u30465 \u20221 ": "\u24191 \u19996 ", "\u22478 \u24066 ": "\u24191 \u24030 \u24066 ", "\u26426 \u26500 ": "\u24191 \u19996 \u30465 \u31246 \u21153 \u23616 ", "\u26469 \u28304 ": "https://guangdong.chinatax.gov.cn/", "\u32423 \u21035 ": "\u30465 \u32423 "\},\
]\
\
TEMPLATES = [\
    \{"\uc0\u30465 \u20221 ": "\u19978 \u28023 ", "\u22478 \u24066 ": "\u19978 \u28023 \u24066 ", "\u27169 \u26495 \u21517 \u31216 ": "\u19978 \u28023 \u24066 \u31038 \u20445 \u30003 \u25253 \u34920 \u65288 \u26376 \u24230 \u65289 ", "\u25253 \u34920 \u31867 \u22411 ": "\u26376 \u24230 \u30003 \u25253 ", "\u29256 \u26412 ": "1.0", "\u24517 \u22635 \u23383 \u27573 ": ["\u21333 \u20301 \u21517 \u31216 ","\u22522 \u25968 "]\},\
    \{"\uc0\u30465 \u20221 ": "\u24191 \u19996 ", "\u22478 \u24066 ": "\u24191 \u24030 \u24066 ", "\u27169 \u26495 \u21517 \u31216 ": "\u24191 \u19996 \u30465 \u31038 \u20445 \u30003 \u25253 \u34920 ", "\u25253 \u34920 \u31867 \u22411 ": "\u26376 \u24230 \u30003 \u25253 ", "\u29256 \u26412 ": "1.0", "\u24517 \u22635 \u23383 \u27573 ": ["\u21333 \u20301 \u21517 \u31216 ","\u22522 \u25968 "]\},\
]\
\
RULES = [\
    \{"\uc0\u30465 \u20221 ": "\u19978 \u28023 ", "\u22478 \u24066 ": "\u19978 \u28023 \u24066 ", "\u25253 \u34920 \u31867 \u22411 ": "\u26376 \u24230 \u30003 \u25253 ", "\u35268 \u21017 ": "\u20859 \u32769 16%/8%", "\u26469 \u28304 \u24341 \u29992 ": "\u27818 \u20154 \u31038 \u35268 \u12308 2024\u12309 22\u21495 "\},\
    \{"\uc0\u30465 \u20221 ": "\u24191 \u19996 ", "\u22478 \u24066 ": "\u24191 \u24030 \u24066 ", "\u25253 \u34920 \u31867 \u22411 ": "\u26376 \u24230 \u30003 \u25253 ", "\u35268 \u21017 ": "\u20859 \u32769 14%/8%", "\u26469 \u28304 \u24341 \u29992 ": "\u31908 \u20154 \u31038 \u35268 \u12308 2024\u12309 8\u21495 "\},\
]\
\
COMPANIES = [\
    \{"\uc0\u20844 \u21496 \u21517 \u31216 ": "\u19978 \u28023 \u31185 \u25216 \u20844 \u21496 ", "\u30465 \u20221 ": "\u19978 \u28023 ", "\u22478 \u24066 ": "\u19978 \u28023 \u24066 ", "\u21306 \u21439 ": "\u28006 \u19996 \u26032 \u21306 "\},\
    \{"\uc0\u20844 \u21496 \u21517 \u31216 ": "\u24191 \u24030 \u31185 \u25216 \u20844 \u21496 ", "\u30465 \u20221 ": "\u24191 \u19996 ", "\u22478 \u24066 ": "\u24191 \u24030 \u24066 ", "\u21306 \u21439 ": "\u22825 \u27827 \u21306 "\},\
]\
\
st.set_page_config(page_title="\uc0\u25253 \u34920 \u21305 \u37197 \u24037 \u20855 ", layout="centered")\
st.title("\uc0\u55357 \u56523  \u31038 \u20445 \u25253 \u34920 \u21305 \u37197 \u24037 \u20855 ")\
st.markdown("---")\
\
province = st.selectbox("\uc0\u30465 \u20221 ", sorted(set(c["\u30465 \u20221 "] for c in COMPANIES)))\
cities = sorted(set(c["\uc0\u22478 \u24066 "] for c in COMPANIES if c["\u30465 \u20221 "] == province))\
city = st.selectbox("\uc0\u22478 \u24066 ", cities)\
districts = sorted(set(c["\uc0\u21306 \u21439 "] for c in COMPANIES if c["\u30465 \u20221 "] == province and c["\u22478 \u24066 "] == city))\
district = st.selectbox("\uc0\u21306 \u21439 ", districts)\
companies = [c["\uc0\u20844 \u21496 \u21517 \u31216 "] for c in COMPANIES if c["\u30465 \u20221 "] == province and c["\u22478 \u24066 "] == city and c["\u21306 \u21439 "] == district]\
company = st.selectbox("\uc0\u20844 \u21496 ", companies)\
report_type = st.selectbox("\uc0\u25253 \u34920 \u31867 \u22411 ", ["\u26376 \u24230 \u30003 \u25253 ", "\u24180 \u24230 \u27719 \u31639 "])\
year = st.selectbox("\uc0\u24180 \u20221 ", [2024, 2025], index=0)\
month = None\
if report_type == "\uc0\u26376 \u24230 \u30003 \u25253 ":\
    month = st.selectbox("\uc0\u26376 \u20221 ", list(range(1,13)), index=11)\
\
if st.button("\uc0\u21305 \u37197 \u23448 \u26041 \u27169 \u26495 "):\
    template = next((t for t in TEMPLATES if t["\uc0\u22478 \u24066 "] == city and t["\u25253 \u34920 \u31867 \u22411 "] == report_type), None)\
    rule = next((r for r in RULES if r["\uc0\u22478 \u24066 "] == city and r["\u25253 \u34920 \u31867 \u22411 "] == report_type), None)\
    if template and rule:\
        st.success("\uc0\u9989  \u21305 \u37197 \u25104 \u21151 \u65281 ")\
        st.write(f"**\uc0\u27169 \u26495 **\u65306 \{template['\u27169 \u26495 \u21517 \u31216 ']\} (v\{template['\u29256 \u26412 ']\})")\
        st.write(f"**\uc0\u35268 \u21017 **\u65306 \{rule['\u35268 \u21017 ']\}")\
        st.write(f"**\uc0\u26469 \u28304 **\u65306 [\{rule['\u26469 \u28304 \u24341 \u29992 ']\}](https://shanghai.chinatax.gov.cn/)")\
        \
        if st.button("\uc0\u29983 \u25104 \u20551 \u25968 \u25454 Excel\u65288 \u24453 \u22797 \u26680 \u29256 \u65289 "):\
            wb = Workbook()\
            ws = wb.active\
            ws.append(['\uc0\u20844 \u21496 ', '\u38505 \u31181 ', '\u22522 \u25968 ', '\u21333 \u20301 \u37329 \u39069 ', '\u20010 \u20154 \u37329 \u39069 '])\
            ws.append([company, '\uc0\u20859 \u32769 \u20445 \u38505 ', 8000, 1280, 640])\
            ws.insert_rows(1)\
            ws['A1'] = '\uc0\u9888 \u65039  \u24453 \u22797 \u26680 \u29256 '\
            ws.merge_cells('A1:E1')\
            audit = wb.create_sheet('AuditTrail')\
            audit.append(['\uc0\u26102 \u38388 ', '\u25805 \u20316 '])\
            audit.append([datetime.now().isoformat(), '\uc0\u29983 \u25104 '])\
            output = BytesIO()\
            wb.save(output)\
            output.seek(0)\
            st.download_button("\uc0\u19979 \u36733 Excel", data=output, file_name=f"\{company\}_\u24453 \u22797 \u26680 .xlsx")\
    else:\
        st.error("\uc0\u26410 \u21305 \u37197 \u21040 \u27169 \u26495 ")\
\
with st.expander("\uc0\u26597 \u30475 \u25152 \u26377 \u25968 \u25454 "):\
    st.dataframe(pd.DataFrame(SOURCES))\
    st.dataframe(pd.DataFrame(TEMPLATES))\
    st.dataframe(pd.DataFrame(RULES))}
